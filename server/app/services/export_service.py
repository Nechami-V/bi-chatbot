"""
Export Service - Generate Excel/CSV exports from natural language queries

This service handles exporting data to Excel and CSV formats.
Uses the Exports.yaml prompt to generate SQL queries optimized for export.

Author: BI Chatbot Team
Version: 1.0.0
"""

import csv
import io
import logging
from typing import Dict, Optional, Any, List, Literal
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.services.ai_service import AIService
from app.services.sql_validator import validate_sql, SQLValidationError

# Configure logger
logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting data to Excel/CSV based on natural language queries
    """

    def __init__(self, db: Session, client_id: str = "KT"):
        self.db = db
        self.client_id = client_id
        self.ai_service = AIService(db, client_id=client_id)

    def _get_schema_brief(self) -> str:
        """Build a brief schema description from AI service schema_info"""
        schema_info = self.ai_service.schema_info
        lines = []
        for table_name, meta in schema_info.get("tables", {}).items():
            columns = [f"{col['name']} ({col['type']})" for col in meta.get("columns", [])]
            pk = meta.get("primary_key", "")
            lines.append(f"Table {table_name}: columns={', '.join(columns)}, PK={pk}")
        return "\n".join(lines)

    def _get_ontology_section(self) -> str:
        """Build ontology description from config_loader"""
        from app.config_loader import config_loader
        try:
            ontology = config_loader.load_shared_ontology()
            lines = []
            for entity_name, entity in ontology.entities.items():
                lines.append(f"Entity {entity_name}: table={entity.physical_table}, key={entity.key_field}")
            return "\n".join(lines)
        except Exception:
            return "No ontology available"

    def _get_dialect(self) -> str:
        """Get database dialect from datasource config"""
        from app.config_loader import config_loader
        try:
            datasource, _ = config_loader.load_client_config(self.client_id)
            # Handle both enum and string
            if hasattr(datasource.database_type, 'value'):
                db_type = datasource.database_type.value.lower()
            else:
                db_type = str(datasource.database_type).lower()
            
            # Map to standard dialect names used in prompts
            dialect_map = {
                'sqlite': 'SQLite',
                'sqlserver': 'SQLServer',
                'mssql': 'SQLServer',
                'postgresql': 'Postgres',
                'postgres': 'Postgres',
                'mysql': 'MySQL'
            }
            
            dialect = dialect_map.get(db_type, 'SQLServer')
            logger.info(f"Detected database dialect: {db_type} -> {dialect}")
            return dialect
        except Exception as e:
            logger.warning(f"Could not determine dialect from datasource: {e}")
            return 'SQLite'

    def _convert_limit_to_top(self, sql: str) -> str:
        """
        Convert MySQL/PostgreSQL LIMIT clause to SQL Server TOP clause.
        
        Examples:
        - SELECT * FROM table LIMIT 10 → SELECT TOP 10 * FROM table
        - SELECT id, name FROM users ORDER BY id LIMIT 5 → SELECT TOP 5 id, name FROM users ORDER BY id
        """
        import re
        
        # Pattern: LIMIT followed by a number at the end of query
        pattern = r'\bLIMIT\s+(\d+)\s*;?\s*$'
        match = re.search(pattern, sql, re.IGNORECASE)
        
        if match:
            limit_num = match.group(1)
            # Remove the LIMIT clause
            sql_without_limit = re.sub(pattern, '', sql, flags=re.IGNORECASE).strip()
            
            # Add TOP after SELECT
            sql_with_top = re.sub(
                r'\bSELECT\b',
                f'SELECT TOP {limit_num}',
                sql_without_limit,
                count=1,
                flags=re.IGNORECASE
            )
            
            logger.info(f"Converted LIMIT {limit_num} to TOP {limit_num}")
            return sql_with_top
        
        return sql

    def _translate_to_sql_server(self, sql: str) -> str:
        """
        Translate SQLite syntax to SQL Server T-SQL
        
        Args:
            sql: SQL query that might contain SQLite syntax
            
        Returns:
            SQL query with T-SQL syntax
        """
        import re
        
        logger.info(f"Translating SQL (before): {sql[:200]}...")
        
        # Pattern 1: strftime('%Y-%m', column) → FORMAT(column, 'yyyy-MM')
        # This catches both SELECT and ORDER BY cases
        sql = re.sub(
            r"strftime\(['\"]%Y-%m['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"FORMAT(\1, 'yyyy-MM')",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 2: strftime('%Y-%m-%d', column) → CAST(column AS DATE)
        sql = re.sub(
            r"strftime\(['\"]%Y-%m-%d['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"CAST(\1 AS DATE)",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 3: strftime('%Y', column) → YEAR(column)
        sql = re.sub(
            r"strftime\(['\"]%Y['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"YEAR(\1)",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 4: strftime('%m', column) → MONTH(column)
        sql = re.sub(
            r"strftime\(['\"]%m['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"MONTH(\1)",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 5: strftime('%W', column) → DATEPART(week, column)
        sql = re.sub(
            r"strftime\(['\"]%W['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"DATEPART(week, \1)",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 6: strftime('%d', column) → DAY(column)
        sql = re.sub(
            r"strftime\(['\"]%d['\"]\s*,\s*([a-zA-Z_][a-zA-Z0-9_.]+)\)",
            r"DAY(\1)",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 7: date('now') → GETDATE()
        sql = re.sub(
            r"date\(['\"]now['\"]\)",
            r"GETDATE()",
            sql,
            flags=re.IGNORECASE
        )
        
        # Pattern 8: LIMIT N → TOP N (convert properly)
        limit_match = re.search(r"\s+LIMIT\s+(\d+)\s*;?\s*$", sql, flags=re.IGNORECASE)
        if limit_match:
            limit_num = limit_match.group(1)
            # Remove LIMIT clause
            sql = re.sub(r"\s+LIMIT\s+\d+\s*;?\s*$", "", sql, flags=re.IGNORECASE)
            # Add TOP after SELECT
            sql = re.sub(
                r"\bSELECT\b",
                f"SELECT TOP {limit_num}",
                sql,
                count=1,
                flags=re.IGNORECASE
            )
            logger.info(f"Converted LIMIT {limit_num} to TOP {limit_num}")
        
        logger.info(f"Translated SQL (after): {sql[:200]}...")
        return sql

    def generate_export_sql(
        self,
        question: str,
        granularity_hint: Optional[str] = None,
        limit_hint: Optional[int] = None,
        date_range_hint: Optional[str] = None
    ) -> str:
        """
        Generate SQL query for export using the Exports.yaml prompt
        
        Args:
            question: Natural language question in Hebrew
            granularity_hint: Optional hint for export granularity (e.g., "Customer", "Order")
            limit_hint: Optional limit for number of rows
            date_range_hint: Optional date range filter
            
        Returns:
            SQL query string optimized for export
        """
        try:
            # Get schema brief and ontology
            schema_brief = self._get_schema_brief()
            ontology = self._get_ontology_section()
            dialect = self._get_dialect()

            # Load Exports.yaml prompt template
            from app.simple_config import config
            from pathlib import Path
            import yaml
            
            prompt_file = Path(config.PROMPTS_DIR) / config.PACK / "Exports.yaml"
            if not prompt_file.exists():
                raise FileNotFoundError(f"Exports.yaml not found at {prompt_file}")
            
            with open(prompt_file, 'r', encoding='utf-8') as f:
                prompt_data = yaml.safe_load(f)
            
            # Get the prompt template
            prompt_template = prompt_data.get('prompt', '')
            
            # Replace placeholders (both LIMIT_HINT and LIMIT_OR_EMPTY are used in template)
            limit_value = str(limit_hint) if limit_hint else ""
            date_value = date_range_hint or ""
            granularity_value = granularity_hint or ""
            
            export_prompt = prompt_template.format(
                DIALECT=dialect,
                SCHEMA_BRIEF=schema_brief,
                ONTOLOGY=ontology,
                QUESTION=question,
                GRANULARITY_HINT_OR_EMPTY=granularity_value,
                LIMIT_OR_EMPTY=limit_value,
                LIMIT_HINT=limit_value,  # Also used in the template
                DATE_RANGE_OR_EMPTY=date_value,
                DATE_RANGE_HINT=date_value  # Also used in the template
            )

            logger.info(f"Generating export SQL for question: {question}")

            # Call OpenAI to generate SQL
            from openai import OpenAI
            from app.simple_config import config
            
            client = OpenAI(api_key=config.OPENAI_API_KEY)
            response = client.chat.completions.create(
                model=config.OPENAI_MODEL,
                messages=[{"role": "user", "content": export_prompt}],
                temperature=0.0,
                max_tokens=1000
            )

            sql_query = response.choices[0].message.content.strip()
            
            # Clean up the SQL (remove markdown code blocks if present)
            sql_query = sql_query.replace("```sql", "").replace("```", "").strip()
            
            # Auto-fix LIMIT to TOP before validation
            if "LIMIT" in sql_query.upper():
                logger.warning("⚠️ Found LIMIT clause in export SQL - auto-converting to TOP")
                sql_query = self._convert_limit_to_top(sql_query)
            
            # ⚡ VALIDATE SQL SYNTAX BEFORE TRANSLATION
            try:
                validate_sql(sql_query, strict=True)
                logger.info("✅ Export SQL validation passed - using only SQL Server syntax")
            except SQLValidationError as ve:
                error_msg = f"Generated export SQL contains forbidden syntax:\n{ve.message}\n\nGenerated SQL:\n{sql_query}"
                logger.error(f"❌ Export SQL validation failed: {error_msg}")
                raise ValueError(error_msg)
            
            # Apply SQL Server translation if needed (as safety net)
            if dialect == "SQLServer":
                sql_query = self._translate_to_sql_server(sql_query)
            
            logger.info(f"Generated export SQL: {sql_query}")
            return sql_query

        except Exception as e:
            logger.error(f"Error generating export SQL: {str(e)}")
            raise

    def execute_export_query(self, sql_query: str) -> List[Dict[str, Any]]:
        """
        Execute SQL query and return results as list of dictionaries
        
        Args:
            sql_query: SQL query to execute
            
        Returns:
            List of row dictionaries
        """
        try:
            # Auto-fix LIMIT to TOP as last resort before execution
            if "LIMIT" in sql_query.upper():
                logger.warning("⚠️ CRITICAL: Found LIMIT in export SQL before execution - emergency fix to TOP")
                sql_query = self._convert_limit_to_top(sql_query)
            
            # Apply SQL translation if needed (for SQL Server)
            dialect = self._get_dialect()
            if dialect == "SQLServer":
                original_sql = sql_query
                sql_query = self._translate_to_sql_server(sql_query)
                if sql_query != original_sql:
                    logger.info(f"Translated export SQL: {original_sql[:100]}... -> {sql_query[:100]}...")
            
            result = self.db.execute(text(sql_query))
            columns = result.keys()
            rows = [dict(zip(columns, row)) for row in result.fetchall()]
            
            logger.info(f"Executed export query, returned {len(rows)} rows")
            return rows

        except Exception as e:
            logger.error(f"Error executing export query: {str(e)}")
            raise

    def create_excel_file(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> bytes:
        """
        Create Excel file from data
        
        Args:
            data: List of row dictionaries
            filename: Optional filename (for logging)
            
        Returns:
            Excel file content as bytes
        """
        if not data:
            raise ValueError("No data to export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # Get column names from first row
        columns = list(data[0].keys())

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = len(str(col_name))
            for row_data in data[:100]:  # Check first 100 rows for performance
                value_length = len(str(row_data.get(col_name, "")))
                max_length = max(max_length, value_length)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Created Excel file with {len(data)} rows, {len(columns)} columns")
        return excel_buffer.getvalue()

    def create_csv_file(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> bytes:
        """
        Create CSV file from data
        
        Args:
            data: List of row dictionaries
            filename: Optional filename (for logging)
            
        Returns:
            CSV file content as bytes
        """
        if not data:
            raise ValueError("No data to export")

        # Create CSV in memory
        csv_buffer = io.StringIO()
        columns = list(data[0].keys())
        
        writer = csv.DictWriter(csv_buffer, fieldnames=columns)
        writer.writeheader()
        writer.writerows(data)

        # Convert to bytes with UTF-8 BOM for Excel compatibility
        csv_content = "\ufeff" + csv_buffer.getvalue()
        
        logger.info(f"Created CSV file with {len(data)} rows, {len(columns)} columns")
        return csv_content.encode('utf-8')

    def export_data(
        self,
        question: str,
        format: Literal["excel", "csv"] = "excel",
        granularity_hint: Optional[str] = None,
        limit_hint: Optional[int] = None,
        date_range_hint: Optional[str] = None
    ) -> tuple[bytes, str]:
        """
        Complete export flow: generate SQL, execute, create file
        
        Args:
            question: Natural language question in Hebrew
            format: Export format ("excel" or "csv")
            granularity_hint: Optional granularity hint
            limit_hint: Optional row limit
            date_range_hint: Optional date range
            
        Returns:
            Tuple of (file_content, filename)
        """
        try:
            # Generate SQL using Exports.yaml
            sql_query = self.generate_export_sql(
                question=question,
                granularity_hint=granularity_hint,
                limit_hint=limit_hint,
                date_range_hint=date_range_hint
            )

            # Execute query
            data = self.execute_export_query(sql_query)

            if not data:
                raise ValueError("השאילתה לא החזירה נתונים")

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if format == "excel":
                file_content = self.create_excel_file(data)
                filename = f"export_{timestamp}.xlsx"
            elif format == "csv":
                file_content = self.create_csv_file(data)
                filename = f"export_{timestamp}.csv"
            else:
                raise ValueError(f"Unsupported format: {format}")

            logger.info(f"Export completed: {filename}, {len(data)} rows")
            return file_content, filename

        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            raise
